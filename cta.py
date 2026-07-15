import argparse
import os
import re
import shutil
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urljoin, urlparse, urlunparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright



NAV_TIMEOUT_MS = 8000
PAGE_LOAD_WAIT_MS = 1200
CLICK_NAV_WAIT_MS = 2500
SPA_ROUTE_WAIT_MS = 500
MAX_DRILLDOWN_DEPTH = 5  

CTA_SELECTOR = "a[href], button, [onclick], [role='button']"

CLOSE_DIALOG_SELECTORS = [
    "[aria-label='Close']", ".modal-close", ".close-button",
    "button:has-text('Close')", "button:has-text('Cancel')",
    "button:has-text('✕')", "button:has-text('×')",
]

ONCLICK_URL_RE = re.compile(
    r"""(?:location\.href|location\.assign|window\.location(?:\.href)?)\s*=?\(?\s*['"]([^'"]+)['"]"""
)




def normalise(url: str) -> str:
    p = urlparse(url)
    path = p.path.rstrip("/") or "/"
    return urlunparse((p.scheme or "https", p.netloc.lower(), path, "", "", ""))


def get_host(url: str) -> str:
    return urlparse(url).netloc.lower().replace("www.", "")




def close_any_open_dialog(page):
    for selector in CLOSE_DIALOG_SELECTORS:
        try:
            el = page.query_selector(selector)
            if el and el.is_visible():
                el.click(timeout=1000)
                return True
        except Exception:
            pass
    try:
        page.keyboard.press("Escape")
    except Exception:
        pass
    return False


def try_href(el, base_url):
    try:
        href = el.get_attribute("href")
    except Exception:
        return None
    if href and href != "#" and not href.startswith("javascript:"):
        return urljoin(base_url, href)
    return None


def try_onclick_regex(el, base_url):
    try:
        onclick = el.get_attribute("onclick")
    except Exception:
        return None
    if not onclick:
        return None
    match = ONCLICK_URL_RE.search(onclick)
    if match:
        return urljoin(base_url, match.group(1))
    return None


def follow_destination(context, url: str):
    """Check where `url` really lands, walking the redirect chain manually so
    we can report the status code of the initial request, any intermediate
    hop(s), and the final destination's status — without opening a real
    browser page.
    """
    chain = []  
    current_url = url

    for _ in range(10):
        try:
            resp = context.request.get(current_url, timeout=NAV_TIMEOUT_MS, max_redirects=0)
        except Exception:
            return None
        status = resp.status
        chain.append((current_url, status))

        if 300 <= status < 400:
            location = resp.headers.get("location")
            if not location:
                break
            current_url = urljoin(current_url, location)
            continue
        break

    if not chain:
        return None

    initial_status = chain[0][1]
    final_url, final_status = chain[-1]
    intermediate_hops = chain[1:-1]  # excludes the initial request and the final landing page
    intermediate_url = " → ".join(f"{u} ({s})" for u, s in intermediate_hops)

    return {
        "final_destination": normalise(final_url),
        "is_redirect": "Y" if len(chain) > 1 else "N",
        "initial_status": initial_status,
        "intermediate_url": intermediate_url,
        "final_status": final_status,
    }





def mark_new_elements(page, uid_counter_start: int) -> int:
    """Tag every not-yet-tagged CTA-matching element with a unique
    data-cta-uid attribute. Returns the updated counter."""
    return page.evaluate(
        """([selector, start]) => {
            let counter = start;
            document.querySelectorAll(selector).forEach(el => {
                if (!el.hasAttribute('data-cta-uid')) {
                    el.setAttribute('data-cta-uid', 'uid-' + (counter++));
                }
            });
            return counter;
        }""",
        [CTA_SELECTOR, uid_counter_start],
    )


def get_marked_uids(page) -> set:
    try:
        return set(page.eval_on_selector_all(
            "[data-cta-uid]", "els => els.map(e => e.getAttribute('data-cta-uid'))"
        ))
    except Exception:
        return set()


class SharedDestinationCache:
    def __init__(self):
        self._data = {}
        self._lock = threading.Lock()

    def get(self, key):
        with self._lock:
            return self._data.get(key)

    def set(self, key, value):
        with self._lock:
            self._data[key] = value

    def __len__(self):
        with self._lock:
            return len(self._data)


class PageResolver:
    """Resolves all CTAs (including nested/drill-down) for a single page."""

    def __init__(self, context, page, source_url, origin_host, destination_cache=None, verbose=True):
        self.context = context
        self.page = page
        self.source_url = source_url
        self.origin_host = origin_host
        self.resolved = []
        self.unresolved = []
        self.uid_counter = 0
        self.visited_uids = set()
        # Shared across ALL pages in the run: avoids re-resolving the same
        # destination (nav/footer/social links repeated on every page) with a
        # fresh browser navigation each time.
        self.destination_cache = destination_cache if destination_cache is not None else SharedDestinationCache()
        self.verbose = verbose

    def run(self):
        self.uid_counter = mark_new_elements(self.page, self.uid_counter)
        processed_keys = set()

        while True:
            uids_now = sorted(get_marked_uids(self.page), key=lambda u: int(u.split("-")[1]))
            name_counts = {}
            next_uid, next_key = None, None

            for uid in uids_now:
                el = self._get_element(uid)
                if el is None:
                    continue
                try:
                    name = (el.inner_text() or "").strip()[:80] or el.get_attribute("aria-label") or "(no label)"
                except Exception:
                    name = "(no label)"
                occurrence = name_counts.get(name, 0)
                name_counts[name] = occurrence + 1
                key = (name, occurrence)
                if key not in processed_keys:
                    next_uid, next_key = uid, key
                    break

            if next_uid is None:
                break  # nothing left unprocessed at the top level

            processed_keys.add(next_key)
            reloaded = self._process_uid(next_uid, breadcrumb=[], depth=0)
            if reloaded:
                # A real page navigation happened and we returned to source_url —
                # every uid marker on the page was wiped by that reload, so the
                # next loop iteration re-scans fresh instead of trusting stale ids.
                # (name, occurrence) pairs stay valid across the reload since the
                # page re-renders identically, so processed_keys is still correct.
                self.uid_counter = mark_new_elements(self.page, self.uid_counter)

    def _get_element(self, uid):
        try:
            return self.page.query_selector(f"[data-cta-uid='{uid}']")
        except Exception:
            return None

    DISMISS_NAMES = {"close", "cancel", "×", "✕", "x"}

    def _process_uid(self, uid, breadcrumb, depth):
        if uid in self.visited_uids:
            return False
        self.visited_uids.add(uid)

        el = self._get_element(uid)
        if el is None:
            return False  # element no longer present (e.g. its parent menu closed)

        try:
            name = (el.inner_text() or "").strip()[:80] or el.get_attribute("aria-label") or "(no label)"
        except Exception:
            name = "(no label)"

        if name.strip().lower() in self.DISMISS_NAMES:
            return False  # dialog-dismissal control, not a real CTA — skip silently

        full_name = " > ".join(breadcrumb + [name]) if breadcrumb else name
        selector_desc = f"[data-cta-uid='{uid}']"

        # Tier 1/2: cheap, no click needed
        dest = try_href(el, self.source_url)
        method = "static_href" if dest else None
        if dest is None:
            dest = try_onclick_regex(el, self.source_url)
            method = "onclick_regex" if dest else None

        if dest is not None:
            self._record_resolved(full_name, selector_desc, dest, method)
            return False

        # Tier 3: click and see what happens
        if depth >= MAX_DRILLDOWN_DEPTH:
            self.unresolved.append({
                "source_url": self.source_url, "cta_name": full_name,
                "reason": f"max drill-down depth ({MAX_DRILLDOWN_DEPTH}) reached",
            })
            return False

        try:
            uids_before = get_marked_uids(self.page)
            url_before = self.page.url

            try:
                with self.page.expect_navigation(timeout=CLICK_NAV_WAIT_MS):
                    el.click(timeout=2000)
                # real navigation happened
                dest_url = self.page.url
                self._record_resolved(full_name, selector_desc, dest_url, "click_simulated")
                self.page.goto(self.source_url, timeout=NAV_TIMEOUT_MS, wait_until="domcontentloaded")
                self.page.wait_for_timeout(PAGE_LOAD_WAIT_MS)
                return True  # signal: full reload happened, uid markers are gone
            except PlaywrightTimeoutError:
                pass

            # no full navigation — check SPA soft route change
            self.page.wait_for_timeout(SPA_ROUTE_WAIT_MS)
            if self.page.url != url_before:
                dest_url = self.page.url
                self._record_resolved(full_name, selector_desc, dest_url, "click_simulated")
                self.page.go_back(timeout=NAV_TIMEOUT_MS)
                self.page.wait_for_timeout(PAGE_LOAD_WAIT_MS)
                return False

            # no navigation at all — did new elements appear? (drill down)
            self.uid_counter = mark_new_elements(self.page, self.uid_counter)
            uids_after = get_marked_uids(self.page)
            new_uids = uids_after - uids_before

            if new_uids:
                before_count = len(self.resolved) + len(self.unresolved)
                for new_uid in sorted(new_uids, key=lambda u: int(u.split("-")[1])):
                    child_reloaded = self._process_uid(new_uid, breadcrumb + [name], depth + 1)
                    if child_reloaded:
                        return True  # bail out of this branch, page already reloaded
                close_any_open_dialog(self.page)  # tidy up before the next sibling

                if len(self.resolved) + len(self.unresolved) == before_count:
                    # every "new" element was just a dismissal control (e.g. a
                    # modal's Close button) — the toggle itself led nowhere,
                    # so log it explicitly rather than letting it vanish.
                    self.unresolved.append({
                        "source_url": self.source_url, "cta_name": full_name,
                        "reason": "opened a dialog/modal with no reachable destination",
                    })
                return False

            closed = close_any_open_dialog(self.page)
            reason = "opened a dialog/modal (closed)" if closed else "no navigation and no new elements (non-navigating control)"
            self.unresolved.append({"source_url": self.source_url, "cta_name": full_name, "reason": reason})
            return False

        except Exception as e:
            self.unresolved.append({
                "source_url": self.source_url, "cta_name": full_name,
                "reason": f"unexpected error: {e}",
            })
            return False

    def _record_resolved(self, name, selector_desc, dest, method):
        cache_key = normalise(dest)
        cached = self.destination_cache.get(cache_key)

        if cached is not None:
            result = cached
            if self.verbose:
                print(f"      [cache hit] {name[:60]} -> {result['final_destination']}")
        else:
            if self.verbose:
                print(f"      resolving: {name[:60]} ...", end=" ", flush=True)
            result = follow_destination(self.context, dest)
            self.destination_cache.set(cache_key, result)
            if self.verbose:
                print("failed" if result is None else "ok")

        if result is None:
            self.unresolved.append({
                "source_url": self.source_url, "cta_name": name,
                "reason": f"resolved to '{dest}' but destination failed to load",
            })
            return

        final = result["final_destination"]
        is_external = "Y" if get_host(final) != self.origin_host else "N"

        self.resolved.append({
            "source_url": self.source_url,
            "cta_name": name,
            "cta_assigned_url": dest,
            "cta_destination": final,
            "is_external": is_external,
            "is_redirect": result["is_redirect"],
            "initial_status": result["initial_status"],
            "intermediate_url": result["intermediate_url"],
            "final_status": result["final_status"],
            "resolution_method": method,
        })


# ----------------------------------------------------------------------------
# Input: read URLs from an Excel file
# ----------------------------------------------------------------------------


def load_urls_from_excel(path: str, column: str | None) -> list[str]:
    wb = load_workbook(path, read_only=True)
    sheet_name = "Included" if "Included" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    col_idx = 0
    if column:
        try:
            col_idx = header.index(column.lower())
        except ValueError:
            col_idx = 0
    elif "url" in header:
        col_idx = header.index("url")

    urls = [row[col_idx] for row in rows[1:] if row and row[col_idx]]
    return urls


# ----------------------------------------------------------------------------
# Output: single color-coded Excel workbook, 3 sheets
# ----------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
ORANGE = PatternFill("solid", fgColor="FCD5B4")
RED = PatternFill("solid", fgColor="FFC7CE")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREY = PatternFill("solid", fgColor="F2F2F2")


def _write_header(ws, headers):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers, max_width=70):
    for col_idx, header in enumerate(headers, start=1):
        best = len(str(header))
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row_cells:
                if cell.value:
                    best = max(best, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(best + 2, max_width)


def write_resolved_sheet(wb, resolved):
    ws = wb.create_sheet("Resolved CTAs")
    headers = ["source_url", "cta_name", "cta_assigned_url", "cta_destination", "is_external", "is_redirect",
               "initial_status", "intermediate_url", "final_status", "resolution_method"]
    _write_header(ws, headers)
    method_fill = {"static_href": GREEN, "onclick_regex": YELLOW, "click_simulated": ORANGE}
    for r in resolved:
        ws.append([r.get(h, "") for h in headers])
        row_num = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.cell(row=row_num, column=headers.index("resolution_method") + 1).fill = method_fill.get(r["resolution_method"], GREY)
        if r["is_external"] == "Y":
            ws.cell(row=row_num, column=headers.index("is_external") + 1).fill = RED
        if r["is_redirect"] == "Y":
            ws.cell(row=row_num, column=headers.index("is_redirect") + 1).fill = YELLOW
        try:
            if int(r["final_status"]) != 200:
                ws.cell(row=row_num, column=headers.index("final_status") + 1).fill = RED
        except (ValueError, TypeError):
            pass
    _autosize(ws, headers)


def write_unresolved_sheet(wb, unresolved):
    ws = wb.create_sheet("Unresolved CTAs")
    headers = ["source_url", "cta_name", "reason"]
    _write_header(ws, headers)
    for r in unresolved:
        ws.append([r.get(h, "") for h in headers])
        row_num = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = RED
            cell.border = THIN_BORDER
    _autosize(ws, headers)


def write_external_domains_sheet(wb, resolved):
    ws = wb.create_sheet("External Domains")
    headers = ["domain", "source", "final_url"]
    _write_header(ws, headers)

    seen_domains = {}
    for r in resolved:
        if r["is_external"] != "Y":
            continue
        domain = get_host(r["cta_destination"])
        if domain not in seen_domains:
            seen_domains[domain] = {"domain": domain, "source": r["source_url"], "final_url": r["cta_destination"]}

    for i, entry in enumerate(sorted(seen_domains.values(), key=lambda d: d["domain"])):
        ws.append([entry[h] for h in headers])
        row_num = ws.max_row
        fill = BLUE if i % 2 == 0 else GREY
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
    _autosize(ws, headers)


def _ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    _write_header(ws, headers)
    return ws


def _append_bordered_row(ws, values):
    ws.append(values)
    row_num = ws.max_row
    for c in range(1, len(values) + 1):
        ws.cell(row=row_num, column=c).border = THIN_BORDER
    return row_num


class ReportWriter:
    RESOLVED_HEADERS = ["source_url", "cta_name", "cta_assigned_url", "cta_destination", "is_external", "is_redirect", "initial_status", "intermediate_url", "final_status", "resolution_method"]
    UNRESOLVED_HEADERS = ["source_url", "cta_name", "reason"]
    EXTERNAL_HEADERS = ["domain", "source", "final_url"]
    PROGRESS_HEADERS = ["source_url", "status", "resolved_count", "unresolved_count", "updated_at"]

    def __init__(self, path, total_urls):
        self.path = path
        self.total_urls = total_urls
        self.lock = threading.Lock()
        self.wb = load_workbook(path) if os.path.exists(path) else Workbook()
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) == 1 and self.wb["Sheet"]["A1"].value is None:
            self.wb.remove(self.wb["Sheet"])
        self.resolved_ws = _ensure_sheet(self.wb, "Resolved CTAs", self.RESOLVED_HEADERS)
        self.unresolved_ws = _ensure_sheet(self.wb, "Unresolved CTAs", self.UNRESOLVED_HEADERS)
        self.external_ws = _ensure_sheet(self.wb, "External Domains", self.EXTERNAL_HEADERS)
        self.progress_ws = _ensure_sheet(self.wb, "Progress", self.PROGRESS_HEADERS)
        self.completed_urls = {r[0] for r in self.progress_ws.iter_rows(min_row=2, values_only=True) if r and r[0] and r[1] == "done"}
        self.external_domains = {r[0] for r in self.external_ws.iter_rows(min_row=2, values_only=True) if r and r[0]}
        self._write_progress_summary()
        self.wb.save(self.path)

    def _write_progress_summary(self):
        done = len(self.completed_urls)
        pct = round((done / self.total_urls) * 100, 2) if self.total_urls else 100.0
        for cell, value in {"G1": "total_urls", "H1": self.total_urls, "G2": "completed_urls", "H2": done, "G3": "percent_complete", "H3": pct}.items():
            self.progress_ws[cell] = value

    def append_page(self, source_url, resolved, unresolved):
        with self.lock:
            for r in resolved:
                row_num = _append_bordered_row(self.resolved_ws, [r.get(h, "") for h in self.RESOLVED_HEADERS])
                self.resolved_ws.cell(row=row_num, column=10).fill = {"static_href": GREEN, "onclick_regex": YELLOW, "click_simulated": ORANGE}.get(r["resolution_method"], GREY)
                if r["is_external"] == "Y":
                    self.resolved_ws.cell(row=row_num, column=5).fill = RED
                if r["is_redirect"] == "Y":
                    self.resolved_ws.cell(row=row_num, column=6).fill = YELLOW
                if str(r.get("final_status", "")) != "200":
                    self.resolved_ws.cell(row=row_num, column=9).fill = RED
                if r["is_external"] == "Y":
                    domain = get_host(r["cta_destination"])
                    if domain not in self.external_domains:
                        self.external_domains.add(domain)
                        ext_row = _append_bordered_row(self.external_ws, [domain, r["source_url"], r["cta_destination"]])
                        fill = BLUE if (self.external_ws.max_row - 2) % 2 == 0 else GREY
                        for c in range(1, 4):
                            self.external_ws.cell(row=ext_row, column=c).fill = fill
            for r in unresolved:
                row_num = _append_bordered_row(self.unresolved_ws, [r.get(h, "") for h in self.UNRESOLVED_HEADERS])
                for c in range(1, 4):
                    self.unresolved_ws.cell(row=row_num, column=c).fill = RED
            self.progress_ws.append([source_url, "done", len(resolved), len(unresolved), datetime.now().isoformat(timespec="seconds")])
            self.completed_urls.add(source_url)
            self._write_progress_summary()
            self.wb.save(self.path)


# ----------------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------------


def process_url(url, origin_host, destination_cache, verbose=False):
    page_start = time.time()
    result = {"url": url, "resolved": [], "unresolved": []}
    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context()
        page = context.new_page()
        try:
            response = page.goto(url, timeout=NAV_TIMEOUT_MS, wait_until="domcontentloaded")
            page.wait_for_timeout(PAGE_LOAD_WAIT_MS)
            if response is None or response.status != 200:
                result["unresolved"].append({"source_url": url, "cta_name": "", "reason": "page failed to load"})
            else:
                resolver = PageResolver(context, page, url, origin_host, destination_cache=destination_cache, verbose=verbose)
                resolver.run()
                result["resolved"] = resolver.resolved
                result["unresolved"] = resolver.unresolved
        except Exception as e:
            result["unresolved"].append({"source_url": url, "cta_name": "", "reason": f"page failed to load: {e}"})
        finally:
            context.close()
            browser.close()
    result["elapsed"] = round(time.time() - page_start, 1)
    return result


def _parse_status(value):
    try:
        if value is None or value == "":
            return None
        return int(str(value).strip())
    except Exception:
        return None


def _backup_file(path):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{path}.bak.{stamp}"
    shutil.copyfile(path, backup_path)
    return backup_path


def _header_index_map(ws):
    header = [c.value for c in ws[1]]
    header_norm = [str(h).strip() if h is not None else "" for h in header]
    return {name: idx + 1 for idx, name in enumerate(header_norm) if name}


def rerun_failed_report(report_path, out_path, status_threshold):
    try:
        wb = load_workbook(report_path)
    except zipfile.BadZipFile:
        print(f"Report workbook is not a valid .xlsx: {report_path}")
        return 1

    if "Resolved CTAs" not in wb.sheetnames:
        print("Report workbook missing 'Resolved CTAs' sheet.")
        return 1

    ws = wb["Resolved CTAs"]
    idx = _header_index_map(ws)
    required = ["source_url", "cta_assigned_url", "cta_destination", "is_external", "is_redirect", "initial_status", "intermediate_url", "final_status"]
    missing = [c for c in required if c not in idx]
    if missing:
        print(f"Report workbook missing columns in 'Resolved CTAs': {missing}")
        return 1

    candidates = []
    for row_num in range(2, ws.max_row + 1):
        final_status = _parse_status(ws.cell(row=row_num, column=idx["final_status"]).value)
        if final_status is None or final_status < status_threshold:
            continue
        assigned = ws.cell(row=row_num, column=idx["cta_assigned_url"]).value
        source = ws.cell(row=row_num, column=idx["source_url"]).value
        if not assigned or not source:
            continue
        candidates.append({
            "row_num": row_num,
            "source_url": str(source).strip(),
            "cta_assigned_url": str(assigned).strip(),
            "old_final_status": final_status,
            "old_destination": ws.cell(row=row_num, column=idx["cta_destination"]).value,
        })

    print(f"\n  Rerun failed mode")
    print(f"  Report       : {report_path}")
    print(f"  Threshold    : {status_threshold}+")
    print(f"  Candidates   : {len(candidates)}\n")
    if not candidates:
        print("  Nothing to rerun.\n")
        return 0

    destination_cache = SharedDestinationCache()
    updated_rows = 0
    updated_urls = set()

    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context()
        try:
            for i, item in enumerate(candidates, start=1):
                result = follow_destination(context, item["cta_assigned_url"])
                if result is None:
                    continue

                final = result["final_destination"]
                source_host = get_host(item["source_url"])
                is_external = "Y" if get_host(final) != source_host else "N"

                new_final_status = _parse_status(result.get("final_status"))
                changed = False

                def _set(col, value):
                    nonlocal changed
                    cell = ws.cell(row=item["row_num"], column=idx[col])
                    old = cell.value
                    cell.value = value
                    if old != value:
                        changed = True

                _set("cta_destination", final)
                _set("is_external", is_external)
                _set("is_redirect", result["is_redirect"])
                _set("initial_status", result["initial_status"])
                _set("intermediate_url", result["intermediate_url"])
                _set("final_status", result["final_status"])

                if is_external == "Y":
                    ws.cell(row=item["row_num"], column=idx["is_external"]).fill = RED
                else:
                    ws.cell(row=item["row_num"], column=idx["is_external"]).fill = GREY

                if result["is_redirect"] == "Y":
                    ws.cell(row=item["row_num"], column=idx["is_redirect"]).fill = YELLOW
                else:
                    ws.cell(row=item["row_num"], column=idx["is_redirect"]).fill = GREY

                if new_final_status is not None and new_final_status != 200:
                    ws.cell(row=item["row_num"], column=idx["final_status"]).fill = RED
                else:
                    ws.cell(row=item["row_num"], column=idx["final_status"]).fill = GREY

                if changed:
                    updated_rows += 1
                    updated_urls.add(item["source_url"])

                pct = (i / len(candidates)) * 100
                print(f"  [{i}/{len(candidates)} | {pct:.1f}%] row {item['row_num']} -> {item['old_final_status']} -> {result['final_status']}")
        finally:
            context.close()
            browser.close()

    if out_path == report_path:
        backup_path = _backup_file(report_path)
        print(f"\n  Backup       : {backup_path}")

    wb.save(out_path)
    print(f"\n  Updated rows : {updated_rows}")
    print(f"  Updated URLs : {len(updated_urls)}")
    print(f"  Output       : {out_path}\n")
    return 0


def main():
    parser = argparse.ArgumentParser(description="Resolve CTA destinations for a list of URLs from an Excel file.")
    parser.add_argument("excel_file", nargs="?", default=None, help="Path to the .xlsx containing URLs")
    parser.add_argument("--column", default=None, help="Column name containing URLs (default: auto-detect 'url')")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N URLs (for testing)")
    parser.add_argument("--out", default="cta_report.xlsx", help="Output workbook path")
    parser.add_argument("--workers", type=int, default=4, help="Pages to process in parallel")
    parser.add_argument("--rerun-report", default=None, help="Path to an existing CTA report .xlsx; reruns rows with final_status >= threshold")
    parser.add_argument("--rerun-threshold", type=int, default=400, help="Status threshold for rerun-report mode (default: 400)")
    args = parser.parse_args()

    if args.rerun_report:
        out_path = args.out or args.rerun_report
        return_code = rerun_failed_report(args.rerun_report, out_path, args.rerun_threshold)
        raise SystemExit(return_code)

    if not args.excel_file:
        print("excel_file is required unless --rerun-report is used")
        raise SystemExit(2)

    all_urls = load_urls_from_excel(args.excel_file, args.column)
    if args.limit:
        all_urls = all_urls[: args.limit]
    if not all_urls:
        print("No URLs found.")
        return

    origin_host = get_host(all_urls[0])
    writer = ReportWriter(args.out, len(all_urls))
    urls = [url for url in all_urls if url not in writer.completed_urls]
    workers = max(1, args.workers)
    print(f"\n  Total pages  : {len(all_urls)}")
    print(f"  Completed    : {len(writer.completed_urls)}")
    print(f"  Remaining    : {len(urls)}")
    print(f"  Workers      : {workers}")
    print(f"  Origin host  : {origin_host}\n")
    if not urls:
        print("  Nothing left to do.\n")
        return

    destination_cache = SharedDestinationCache()
    run_start = time.time()
    done = len(writer.completed_urls)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(process_url, url, origin_host, destination_cache, workers == 1) for url in urls]
        for future in as_completed(futures):
            page_result = future.result()
            writer.append_page(page_result["url"], page_result["resolved"], page_result["unresolved"])
            done += 1
            pct = (done / len(all_urls)) * 100
            elapsed_total = round(time.time() - run_start, 1)
            print(f"  [{done}/{len(all_urls)} | {pct:.1f}%] {page_result['url']} -> {len(page_result['resolved'])} resolved, {len(page_result['unresolved'])} unresolved (cache size: {len(destination_cache)}) [page: {page_result['elapsed']}s | total elapsed: {elapsed_total}s]")

    total_elapsed = time.time() - run_start
    mins, secs = divmod(total_elapsed, 60)
    hrs, mins = divmod(mins, 60)
    print(f"\n  Total: {writer.resolved_ws.max_row - 1} resolved, {writer.unresolved_ws.max_row - 1} unresolved")
    print(f"  Time taken: {int(hrs)}h {int(mins)}m {round(secs, 1)}s  ({round(total_elapsed, 1)}s total)")
    print(f"  Workbook -> {args.out}")
    print("  Done.\n")


if __name__ == "__main__":
    main()